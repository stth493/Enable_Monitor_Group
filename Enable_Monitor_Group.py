from jinja2 import Template
from getpass import getpass
import lxml
import openpyxl
from ncclient import manager, transport, operations

from rpc_monitor_group import *
from restapi import *
from get_evpn_peer import Get_EVPN_Peer

import logging

def make_monitor_group_template(data_node:dict) -> dict:
    return Template(NETCONF_Monitor_Group).render(data_node)

def make_track_interface_template(data_interface:dict) -> dict:
    return Template(NETCONF_track_interface).render(data_interface)

def get_user_credentials() -> tuple:
    user = input('\nUsername: ')
    passwd = getpass()
    return user, passwd

def get_confirmation():
    confirm = input(f'\nPlease confirm these parameters are correct to proceed (Y/N): ')
    return confirm.upper() == 'Y'

class Monitor_Group():
    def __init__(self, user: str, passwd: str) -> None:
        self.user = user
        self.passwd = passwd
        self.CID_list = []
        self.ACM = self.ACM_authentication()
        self.setup_logging()

        if self.ACM is None:
            raise ValueError("Authentication failed. Exiting program.")
        
    def ACM_authentication(self):
        ACM = ACM_API()    
        ACM.set_authenticate(self.user, self.passwd)
        ACM.get_accessToken()
        if not ACM.Isauthenticated: return None 
        return ACM
    
    def setup_logging(self) -> None:
        """
        Set up logging configuration
        """
        logging.basicConfig(
            level=logging.WARNING,
            format="%(asctime)s %(levelname)s %(message)s",
            datefmt="%Y-%m-%d %H:%M:%S",
            filename="Logfile.log"
        )

    def get_node_param(self,ip_node:str) -> dict:
        node_param = {
             'host' : ip_node,
             'port' : '22',
             'username' : self.user,
             'password' : self.passwd,
             'hostkey_verify' : False,
             'device_params' : {'name': "huaweiyang"},
             'allow_agent' : False,
             'look_for_keys' : False,
             'timeout' : 300
         }
        return node_param
    
    def read_node_list(self) -> list:
        IP_list = []
        try:
            workbook = openpyxl.load_workbook('Node.xlsx', data_only=True)
            # Get the specified sheet or the first sheet
            sheet = workbook.active
        except FileNotFoundError as error:
            logging.error(error)
            raise FileNotFoundError(f"Excel file not found in script folder")
        except Exception as error:
            logging.error(error)
            raise Exception(f"Error reading Excel file: {str(error)}")
        
        IP_col = None
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=1, column=col).value == "IP Loopback":
                IP_col = col
        if IP_col is None:
            raise ValueError("CID column not found in the Excel file")
        for row in range(2, sheet.max_row + 1):
            ip = sheet.cell(row=row, column=IP_col).value
            if ip is None: continue  # Only add non-empty values
            IP_list.append(str(ip))
        return IP_list

    """ def read_PE_list(self) -> None:
        data_list = []
        try:
            workbook = openpyxl.load_workbook('Node_Port_Data.xlsx', data_only=True)
            # Get the specified sheet or the first sheet
            sheet = workbook['Profile']
        except FileNotFoundError:
            raise FileNotFoundError(f"Excel file not found in script folder")
        except Exception as e:
            raise Exception(f"Error reading Excel file: {str(e)}")
        
        for col in range(1, sheet.max_column + 1):
            PE = {}
            ip = sheet.cell(row=1, column=col).value.split('/')[1]
            PE_list = []
            for row in range(2,sheet.max_row + 1):
                if sheet.cell(row=row, column=col).value == None:return None
                PE_list.append(sheet.cell(row=row, column=col).value)
            PE['IP'] = ip
            PE['IP_PE_LIST'] = PE_list
            data_list.append(PE)
        workbook.close()
        return data_list
    
    def read_interface_list(self, sheet_name: str = None) -> None:
        data_list = []
        try:
            workbook = openpyxl.load_workbook('Node_Port_Data.xlsx', data_only=True)
            # Get the specified sheet or the first sheet
            sheet = workbook['Interface']
        except FileNotFoundError:
            raise FileNotFoundError(f"Excel file not found in script folder")
        except Exception as e:
            raise Exception(f"Error reading Excel file: {str(e)}")
        
        for col in range(1, sheet.max_column + 1):
            interface = {}
            if sheet.cell(row=1, column=col).value == None:return None
            ip = sheet.cell(row=1, column=col).value.split('/')[1]
            int_list = []
            for row in range(2,sheet.max_row + 1):
                if sheet.cell(row=row, column=col).value == None:continue
                int_list.append(sheet.cell(row=row, column=col).value.replace('(10G)','').replace('(100M)',''))
            interface['IP'] = ip
            interface['INTERFACE_LIST'] = int_list
            data_list.append(interface)
        workbook.close()
        return data_list """

    def NETCONF_config(self, node_param, config_template) -> bool:
        try:
            with manager.connect(**node_param) as node:
                node.edit_config(config_template,target='running',default_operation=None)
                print('\n\nDone! Please re-check the result.\n')        
                return 'Success'      
        except transport.errors.SSHError as e:
            error = f'\nError : Log-in failed : {e}'
        except operations.rpc.RPCError as e:
            error = f'\nError : RPC Config parameters error : {e}'
        except lxml.etree.XMLSyntaxError as e:
            error = f'\n\nError : RPC check-Config XML error : {e}'
        print(error)
        return error
    
    """ def TEST(self):
        self.user = 'satestacc'
        self.passwd = 'ServiceActivation@123'
        IP_AN1 = '172.29.82.5'
        
        PE_list ={
            'IP_PE_LIST' : ['172.29.82.1','172.29.82.3']
        }
        
        interface_list ={
            'INTERFACE_LIST' : ['GigabitEthernet0/1/5', 'GigabitEthernet0/1/6']
        }      
        
        self.Create_profile(IP_AN1, PE_list)
        self.Track_interface(IP_AN1, interface_list) """

    def Create_profile(self, ip_node:str, PE_list:str) -> bool:
        node_param = self.get_node_param(ip_node)
        config_template = make_monitor_group_template(PE_list)
        logging.warning(config_template)
        #print(config_template)
        result = self.NETCONF_config(node_param, config_template)
        if result == 'Success':
            print(result + '\n')
            logging.warning('Successfully configure track profile\n' + str(PE_list))
            return True
        else:
            logging.error('Failed to create track profile : ' + result)
            return False
    
    def Track_interface(self, ip_node:str, interface_list:str) -> bool:
        node_param = self.get_node_param(ip_node)
        config_template = make_track_interface_template(interface_list)
        logging.warning(config_template)
        #print(config_template)
        result = self.NETCONF_config(node_param, config_template)
        if result == 'Success':
            print(result + '\n')
            logging.warning('Successfully track these interfaces\n' + str(interface_list))
            return True
        else:
            logging.error('Failed to track interfaces : ' + result)
            return False
    
    def Get_PE_list(self, ip_loopback:str) -> dict:
        PE_list = {}
        device_params = {
        'device_type': 'huawei',
        'host': ip_loopback, 
        'username': self.user,
        'password': self.passwd,
        }
        get_peer = Get_EVPN_Peer(device_params)
        PE_list['IP_PE_LIST'] = get_peer.get_bgp_evpn_peer()
        get_peer.close_conn()
        return PE_list
    
    def Get_Interface_list(self, ip_loopback:str) -> dict:
        interface_list = {}
        device_params = {
        'device_type': 'huawei',
        'host': ip_loopback, 
        'username': self.user,
        'password': self.passwd,
        }
        get_peer = Get_EVPN_Peer(device_params)
        allint_list = get_peer.get_All_Interface()
        exclude_list = get_peer.get_exclude_interface()
        int_list = [i for i in allint_list if i not in exclude_list]
        interface_list['INTERFACE_LIST'] = int_list
        get_peer.close_conn()
        return interface_list
        
    def run_new(self, ip_loopback:str) -> None:
        PE_dict  = self.Get_PE_list(ip_loopback)
        print('\nHead ring PEs are\n')
        print(PE_dict)
        if get_confirmation():
            self.Create_profile(ip_loopback, PE_dict)
        
        interface_dict = self.Get_Interface_list(ip_loopback)
        print('\n\n')
        for int in interface_dict['INTERFACE_LIST']: print(int)
        if get_confirmation():
            self.Track_interface(ip_loopback, interface_dict)
            
    def run_from_file(self) -> None:
        IP_list = self.read_node_list()
        for ip_loopback in IP_list:
            PE_dict  = self.Get_PE_list(ip_loopback)
            print(PE_dict)             
            self.Create_profile(ip_loopback, PE_dict)
            interface_dict = self.Get_Interface_list(ip_loopback)
            print('\n\n')
            for int in interface_dict['INTERFACE_LIST']: print(int)
            self.Track_interface(ip_loopback, interface_dict)
        
def main() -> None:
    while True:
        user, passwd = get_user_credentials()
        try:
            monitor_group = Monitor_Group(user, passwd)
        except ValueError as e:
            print(f"Error : User/Password incorrect : {e}")
            time.sleep(5)
            return
    
        print("\nPlease choose an option:")
        print("1. Run on a single node")
        print("2. Run as batch from data file")
        print("3. Exit")
        
        choice = input("\nEnter your choice (1-3): ").strip()
        
        if choice == "1":
            ip_loopback = input("IP Node: ")
            print("\nRunning config monitor group...")
            monitor_group.run_new(ip_loopback)
            
        elif choice == "2":
            print("\nRunning from file Node.xlsx...")
            monitor_group.run_from_file()
            
        elif choice == "3":
            print("\nExiting program...")
            break
        else:
            print("\nInvalid choice. Please enter 1, 2, or 3.")
            
if __name__ == "__main__":
    main()